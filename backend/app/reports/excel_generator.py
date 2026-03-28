"""Excel report generator for schema comparison results."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from typing import Any


class ExcelReportGenerator:
    """Generates Excel reports with multiple sheets from schema comparison results."""

    def __init__(self):
        """Initialize Excel report generator."""
        self.wb = Workbook()

        # Style definitions
        self.header_fill = PatternFill(
            start_color="4472C4",
            end_color="4472C4",
            fill_type="solid"
        )
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.header_alignment = Alignment(horizontal="center", vertical="center")

        # Diff type fills
        self.added_fill = PatternFill(
            start_color="C6EFCE",
            end_color="C6EFCE",
            fill_type="solid"
        )  # Green
        self.removed_fill = PatternFill(
            start_color="FFC7CE",
            end_color="FFC7CE",
            fill_type="solid"
        )  # Red
        self.modified_fill = PatternFill(
            start_color="FFEB9C",
            end_color="FFEB9C",
            fill_type="solid"
        )  # Yellow

        # Common borders
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def generate(
        self,
        diff_result: dict[str, Any],
        source_db: str,
        target_db: str,
        output_path: str,
    ) -> str:
        """Generate Excel report with multiple sheets.

        Args:
            diff_result: SchemaDiffResponse dict with comparison results
            source_db: Source database name/identifier
            target_db: Target database name/identifier
            output_path: Path to write .xlsx file

        Returns:
            Path to generated Excel file
        """
        # Remove default sheet
        default_sheet = self.wb.active
        if default_sheet:
            self.wb.remove(default_sheet)

        # Create sheets
        self._create_summary_sheet(diff_result, source_db, target_db)
        self._create_diff_sheet(
            "Column Differences",
            diff_result.get("column_diffs", []),
            ["Column", "Diff Type", "Source", "Target", "Differences"],
            has_source_target=True
        )
        self._create_diff_sheet(
            "Index Differences",
            diff_result.get("index_diffs", []),
            ["Index", "Diff Type", "Source", "Target", "Differences"],
            has_source_target=True
        )
        self._create_diff_sheet(
            "Constraint Differences",
            diff_result.get("constraint_diffs", []),
            ["Constraint", "Type", "Diff Type", "Source", "Target", "Differences"],
            has_constraint_type=True
        )

        self.wb.save(output_path)
        return output_path

    def _create_summary_sheet(
        self,
        diff_result: dict[str, Any],
        source_db: str,
        target_db: str,
    ) -> None:
        """Create summary sheet with overview and statistics."""
        ws = self.wb.create_sheet("Summary")

        # Title
        ws["A1"] = "Database Schema Comparison Report"
        ws["A1"].font = Font(bold=True, size=16)

        # Metadata
        ws["A3"] = "Generated:"
        ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws["A4"] = "Source Database:"
        ws["B4"] = source_db
        ws["A5"] = "Target Database:"
        ws["B5"] = target_db

        # Statistics header
        ws["A7"] = "Statistics"
        ws["A7"].font = Font(bold=True, size=12)
        ws["A7"].border = self.thin_border

        # Statistics data
        column_diffs = diff_result.get("column_diffs", [])
        index_diffs = diff_result.get("index_diffs", [])
        constraint_diffs = diff_result.get("constraint_diffs", [])

        stats = [
            ("Column Differences", len(column_diffs)),
            ("Index Differences", len(index_diffs)),
            ("Constraint Differences", len(constraint_diffs)),
            ("Total Differences", len(column_diffs) + len(index_diffs) + len(constraint_diffs)),
        ]

        row = 8
        for label, count in stats:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = count
            ws[f"A{row}"].border = self.thin_border
            ws[f"B{row}"].border = self.thin_border
            row += 1

        # Breakdown by diff type
        row += 1
        ws[f"A{row}"] = "Breakdown by Type"
        ws[f"A{row}"].font = Font(bold=True, size=11)

        row += 1
        ws[f"A{row}"] = "Type"
        ws[f"B{row}"] = "Columns"
        ws[f"C{row}"] = "Indexes"
        ws[f"D{row}"] = "Constraints"
        for cell in ws[f"A{row}:D{row}"]:
            for c in cell:
                c.font = self.header_font
                c.fill = self.header_fill
                c.alignment = self.header_alignment
                c.border = self.thin_border

        # Count by type
        def count_by_type(diffs: list) -> dict[str, int]:
            counts = {"added": 0, "removed": 0, "modified": 0}
            for d in diffs:
                diff_type = d.get("diff_type", "modified")
                if diff_type in counts:
                    counts[diff_type] += 1
            return counts

        column_counts = count_by_type(column_diffs)
        index_counts = count_by_type(index_diffs)
        constraint_counts = count_by_type(constraint_diffs)

        for diff_type in ["added", "removed", "modified"]:
            row += 1
            ws[f"A{row}"] = diff_type.upper()
            ws[f"B{row}"] = column_counts.get(diff_type, 0)
            ws[f"C{row}"] = index_counts.get(diff_type, 0)
            ws[f"D{row}"] = constraint_counts.get(diff_type, 0)

            # Apply color fill based on diff type
            fill_color = {
                "added": self.added_fill,
                "removed": self.removed_fill,
                "modified": self.modified_fill,
            }.get(diff_type)

            for cell in ws[f"A{row}:D{row}"]:
                for c in cell:
                    c.border = self.thin_border
                    if fill_color:
                        c.fill = fill_color

        # Column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 15

    def _create_diff_sheet(
        self,
        sheet_name: str,
        diffs: list[dict],
        headers: list[str],
        has_source_target: bool = False,
        has_constraint_type: bool = False,
    ) -> None:
        """Create a difference sheet with specified columns.

        Args:
            sheet_name: Name of the sheet
            diffs: List of diff dictionaries
            headers: Column headers
            has_source_target: If True, expects source_definition/target_definition keys
            has_constraint_type: If True, includes constraint_type column
        """
        ws = self.wb.create_sheet(sheet_name)

        # Header row
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border

        # Data rows
        for row_idx, diff in enumerate(diffs, start=2):
            row_data = self._format_diff_row(
                diff,
                has_source_target=has_source_target,
                has_constraint_type=has_constraint_type,
            )

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

                # Apply diff type fill
                diff_type = diff.get("diff_type", "modified")
                fill_color = {
                    "added": self.added_fill,
                    "removed": self.removed_fill,
                    "modified": self.modified_fill,
                }.get(diff_type)
                if fill_color:
                    cell.fill = fill_color

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

    def _format_diff_row(
        self,
        diff: dict,
        has_source_target: bool = False,
        has_constraint_type: bool = False,
    ) -> list[str]:
        """Format a diff dictionary into a row of values."""
        row = []

        # Name column (column name, index name, or constraint name)
        if "column_name" in diff:
            row.append(diff["column_name"])
        elif "index_name" in diff:
            row.append(diff["index_name"])
        elif "constraint_name" in diff:
            row.append(diff["constraint_name"])

        # Constraint type (if applicable)
        if has_constraint_type:
            row.append(diff.get("constraint_type", ""))

        # Diff type
        row.append(diff.get("diff_type", "").upper())

        # Source and Target definitions
        if has_source_target or has_constraint_type:
            source_def = diff.get("source_definition")
            target_def = diff.get("target_definition")

            # Format source
            if source_def:
                if "columns" in source_def:
                    source_str = ", ".join(str(c) for c in source_def.get("columns", []))
                elif "type" in source_def:
                    source_str = f"{'NOT NULL ' if not source_def.get('nullable', True) else ''}{source_def.get('type', '')}"
                else:
                    source_str = str(source_def)
            else:
                source_str = "N/A"
            row.append(source_str)

            # Format target
            if target_def:
                if "columns" in target_def:
                    target_str = ", ".join(str(c) for c in target_def.get("columns", []))
                elif "type" in target_def:
                    target_str = f"{'NOT NULL ' if not target_def.get('nullable', True) else ''}{target_def.get('type', '')}"
                else:
                    target_str = str(target_def)
            else:
                target_str = "N/A"
            row.append(target_str)

        # Differences list
        differences = diff.get("differences", [])
        row.append("\n".join(differences) if differences else "")

        return row
